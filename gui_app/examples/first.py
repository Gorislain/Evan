import credentials as credentials
import matplotlib.pyplot as plt
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter as tk
from tkinter import *
from tkinter import ttk
import pandas as pd
import openpyxl
import math
from tkinter import ALL, EventType
import os
from openpyxl import Workbook, load_workbook

cred = credentials.Certificate(
    r'C:\...json')
firebase_admin.initialize_app(cred, {'databaseURL': 'https:...firebasedatabase.app/'})
workbook = openpyxl.Workbook()

def firebase(dart):
    VPanel = []
    IPanel = []
    PPanel = []
    VBatt = []
    IBatt = []
    PBatt = []
    VLoad = []
    ILoad = []
    PLoad = []
    full = []
    full_ex = []
    time = []
    date = []
    ate = 0
    rate = 0
    fall = 0
    bit = 0
    b1 = 0
    b2 = 0
    b3 = 0
    b4 = 0
    b5 = 0
    b6 = 0
    b7 = 0
    b8 = 0
    b9 = 0
    a1 = ''
    a2 = ''
    a3 = ''
    a4 = ''
    a5 = ''
    a6 = ''
    a7 = ''
    a8 = ''
    a9 = ''

    ref = db.reference('/01/' + str(dart) + ':')  # Replace 'your_data_path' with the desired path in the database
    data = ref.get()
    if data is not None:
        print("Retrieved data:")
        for key, value in data.items():
            print("Key:", key)
            print("Value:", value)
            print()
            if key == 'VPanel':
                b1 = value
            elif key == 'IPanel':
                b2 = value
            elif key == 'PPanel':
                b3 = value
            elif key == 'VBatt':
                b4 = value
            elif key == 'IBatt':
                b5 = value
            elif key == 'PBatt':
                b6 = value
            elif key == 'VLoad':
                b7 = value
            elif key == 'ILoad':
                b8 = value
            elif key == 'PLoad':
                b9 = value
            else:
                f = 0
        values_list1 = b1.split('_')
        values_list2 = b2.split('_')
        values_list3 = b3.split('_')
        values_list4 = b4.split('_')
        values_list5 = b5.split('_')
        values_list6 = b6.split('_')
        values_list7 = b7.split('_')
        values_list8 = b8.split('_')
        values_list9 = b9.split('_')
        lis = []
        for vp in range(660):
            lis.append(float(values_list1[vp]))
        VPanel = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list2[vp]))
        IPanel = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list3[vp]))
        PPanel = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list4[vp]))
        VBatt = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list5[vp]))
        IBatt = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list6[vp]))
        PBatt = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list7[vp]))
        VLoad = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list8[vp]))
        ILoad = lis
        lis = []
        for vp in range(660):
            lis.append(float(values_list9[vp]))
        PLoad = lis
        for hour in range(8, 19):
            for minute in range(60):
                date.append('/' + str(dart) + ':' + str(hour) + ':' + str(minute))
                time.append(str(hour) + ':' + str(minute))
                rate += 1
        return VPanel, IPanel, PPanel, VBatt, IBatt, PBatt, VLoad, ILoad, PLoad, date, dart


    else:
        for hour in range(8, 19):
            for minute in range(60):
                date.append('/' + str(dart) + ':' + str(hour) + ':' + str(minute))
                time.append(str(hour) + ':' + str(minute))
                rate += 1
        timeset = 60
        chec = 0
        for ui in range(11):
            opiu = timeset*ui
            ref = db.reference(date[opiu])  # Replace 'your_data_path' with the desired path in the database
            data = ref.get()
            if data is not None:
                chec += 1
            else:
                chec += 0
        if chec > 0:
            print("No data found in the specified path.")
            for hour in range(8, 19):
                for minute in range(60):
                    date.append('/' + str(dart) + ':' + str(hour) + ':' + str(minute))
                    time.append(str(hour) + ':' + str(minute))
                    rate += 1
            print(date)

            for i in range(rate):
                ref = db.reference(date[i])
                data = ref.get()
                if data is not None:
                    print(date[i] + ": Значение:", data)
                    values_list = data.split('_')
                    print(date[i] + ": Разделенные значения:", values_list)
                    if fall == 0:
                        if 11.4 > float(values_list[3]) or float(values_list[3]) > 12.1:
                            fall += 1
                        else:
                            fall = 0
                    else:
                        fall = 1
                    if fall == 0:
                        if float(values_list[1]) >= 0.5 or float(values_list[4]) >= 0.5 or float(values_list[7]) >= 0.5:
                            fall += 1
                        else:
                            fall = 0
                    else:
                        fall = 1
                    if fall == 0:
                        if float(values_list[2]) >= 6 or float(values_list[5]) >= 6 or float(values_list[8]) >= 6:
                            fall += 1
                        else:
                            fall = 0
                    else:
                        fall = 1
                    if fall == 0:
                        if 11.4 >= float(values_list[6]):
                            fall += 1
                        else:
                            fall = 0
                    else:
                        fall = 1
                    if fall == 0:
                        ate += 1
                        bit = 0
                        VPanel.append(values_list[0])
                        IPanel.append(values_list[1])
                        PPanel.append(values_list[2])
                        VBatt.append(values_list[3])
                        IBatt.append(values_list[4])
                        PBatt.append(values_list[5])
                        VLoad.append(values_list[6])
                        ILoad.append(values_list[7])
                        PLoad.append(values_list[8])
                    else:
                        op = 0
                        ap = 0
                        kp = 0
                        sar1 = 0
                        sar2 = 0
                        sar3 = 0
                        for ki1 in values_list:
                            if sar1 >= 6:
                                sar1 = sar3 - 9
                            else:
                                sar1 = sar3

                            if sar2 >= 3:
                                sar2 = sar3 - 9
                            else:
                                sar2 = sar3

                            if float(values_list[3 + sar1]) == float(values_list[6 + sar2]) and 11.4 < float(
                                    values_list[3 + sar1]) < 12.1:
                                kp += 1
                                sar3 += 1
                            else:
                                kp += 0
                        for ki in values_list:
                            if float(ki) == 0.00:
                                op += 1
                            elif float(ki) > 18:
                                ap += 1
                            else:
                                op += 0
                        yt = True
                        to = 0
                        if op > 4 or kp == 0:
                            VPanel.append(VPanel[ate - 1])
                            IPanel.append(IPanel[ate - 1])
                            PPanel.append(PPanel[ate - 1])
                            VBatt.append(VBatt[ate - 1])
                            IBatt.append(IBatt[ate - 1])
                            PBatt.append(PBatt[ate - 1])
                            VLoad.append(VLoad[ate - 1])
                            ILoad.append(ILoad[ate - 1])
                            PLoad.append(PLoad[ate - 1])
                            bit = 0
                            ate += 1
                            fall = 0
                        elif ap >= 1:
                            while yt:
                                if float(values_list[to]) <= 18.00:
                                    to += 1
                                else:
                                    yt = False
                            if to == 0:
                                values_list[to] = VPanel[ate - 1]
                            elif to == 1:
                                values_list[to] = IPanel[ate - 1]
                            elif to == 2:
                                values_list[to] = float(values_list[0]) * float(values_list[1])
                            elif to == 3:
                                values_list[to] = VBatt[ate - 1]
                            elif to == 4:
                                values_list[to] = IBatt[ate - 1]
                            elif to == 5:
                                values_list[to] = float(values_list[3]) * float(values_list[4])
                            elif to == 6:
                                values_list[to] = VLoad[ate - 1]
                            elif to == 7:
                                values_list[to] = ILoad[ate - 1]
                            else:
                                values_list[to] = float(values_list[6]) * float(values_list[7])
                            VPanel.append(values_list[0])
                            IPanel.append(values_list[1])
                            PPanel.append(values_list[2])
                            VBatt.append(values_list[3])
                            IBatt.append(values_list[4])
                            PBatt.append(values_list[5])
                            VLoad.append(values_list[6])
                            ILoad.append(values_list[7])
                            PLoad.append(values_list[8])
                            bit = 0
                            ate += 1
                            fall = 0

                        else:
                            position = 0
                            for g in range(9):
                                if g > 2:
                                    ur = -9 + g
                                else:
                                    ur = g
                                if g > 5:
                                    ur1 = -9 + g
                                else:
                                    ur1 = g
                                equels = float(values_list[3 + ur1]) == float(values_list[6 + ur]) and float(
                                    values_list[3 + ur1]) > 11.4
                                while equels:
                                    position = g
                                    equels = False
                            for t in range(position):
                                go = []
                                go.append(values_list[1])
                                go.append(values_list[2])
                                go.append(values_list[3])
                                go.append(values_list[4])
                                go.append(values_list[5])
                                go.append(values_list[6])
                                go.append(values_list[7])
                                go.append(values_list[8])
                                go.append(values_list[0])
                                values_list = go
                            if float(values_list[3]) != float(values_list[6]) and 11.4 < float(values_list[3]) < 12.1:
                                agr = 0
                                agr1 = 0
                                agr2 = 0
                                arg1 = 0
                                arg2 = 0
                                flag = True
                                while flag:
                                    if agr1 >= 6:
                                        arg1 = -9 + agr1
                                    else:
                                        arg1 = agr1
                                    if agr2 >= 3:
                                        arg2 = -9 + agr2
                                    else:
                                        arg2 = agr2
                                    if float(values_list[3 + arg1]) != float(values_list[6 + arg2]):
                                        agr += 1
                                        agr1 += 1
                                        agr2 += 1
                                    else:
                                        for ig in range(agr):
                                            dop = []
                                            dop.append(values_list[1])
                                            dop.append(values_list[2])
                                            dop.append(values_list[3])
                                            dop.append(values_list[4])
                                            dop.append(values_list[5])
                                            dop.append(values_list[6])
                                            dop.append(values_list[7])
                                            dop.append(values_list[8])
                                            dop.append(values_list[9])
                                            values_list = dop
                                            print(dop)
                                        bit = 0
                                        ate += 1
                                        VPanel.append(values_list[0])
                                        IPanel.append(values_list[1])
                                        PPanel.append(values_list[2])
                                        VBatt.append(values_list[3])
                                        IBatt.append(values_list[4])
                                        PBatt.append(values_list[5])
                                        VLoad.append(values_list[6])
                                        ILoad.append(values_list[7])
                                        PLoad.append(values_list[8])
                                        fall = 0
                            else:
                                bit = 0
                                ate += 1
                                VPanel.append(values_list[0])
                                IPanel.append(values_list[1])
                                PPanel.append(values_list[2])
                                VBatt.append(values_list[3])
                                IBatt.append(values_list[4])
                                PBatt.append(values_list[5])
                                VLoad.append(values_list[6])
                                ILoad.append(values_list[7])
                                PLoad.append(values_list[8])
                                fall = 0


                else:
                    if len(VPanel) > 1:
                        if bit <= 9:
                            ate += 1
                            print(date[i] + ": Данные отсутствуют.")
                            VPanel.append(VPanel[ate - 2])
                            IPanel.append(IPanel[ate - 2])
                            PPanel.append(PPanel[ate - 2])
                            VBatt.append(VBatt[ate - 2])
                            IBatt.append(IBatt[ate - 2])
                            PBatt.append(PBatt[ate - 2])
                            VLoad.append(VLoad[ate - 2])
                            ILoad.append(ILoad[ate - 2])
                            PLoad.append(PLoad[ate - 2])
                            bit += 1
                        else:
                            ate += 1
                            VPanel.append(0)
                            IPanel.append(0)
                            PPanel.append(0)
                            VBatt.append(0)
                            IBatt.append(0)
                            PBatt.append(0)
                            VLoad.append(0)
                            ILoad.append(0)
                            PLoad.append(0)
                    else:
                        ate += 1
                        print(date[i] + ": Данные отсутствуют.")
                        VPanel.append(0)
                        IPanel.append(0)
                        PPanel.append(0)
                        VBatt.append(0)
                        IBatt.append(0)
                        PBatt.append(0)
                        VLoad.append(0)
                        ILoad.append(0)
                        PLoad.append(0)

            print(ate)
            full.append(VPanel)
            full.append(IPanel)
            full.append(PPanel)
            full.append(VBatt)
            full.append(IBatt)
            full.append(PBatt)
            full.append(VLoad)
            full.append(ILoad)
            full.append(PLoad)

            lis = []
            tic = 0
            for vp in VPanel:
                lis.append(float(VPanel[tic]))
                tic += 1
            VPanel = lis
            lis = []
            tic = 0
            for vp in IPanel:
                lis.append(float(IPanel[tic]))
                tic += 1
            IPanel = lis
            lis = []
            tic = 0
            for vp in PPanel:
                lis.append(float(PPanel[tic]))
                tic += 1
            PPanel = lis
            lis = []
            tic = 0
            for vp in VBatt:
                lis.append(float(VBatt[tic]))
                tic += 1
            VBatt = lis
            lis = []
            tic = 0
            for vp in IBatt:
                lis.append(float(IBatt[tic]))
                tic += 1
            IBatt = lis
            lis = []
            tic = 0
            for vp in PBatt:
                lis.append(float(PBatt[tic]))
                tic += 1
            PBatt = lis
            lis = []
            tic = 0
            for vp in VLoad:
                lis.append(float(VLoad[tic]))
                tic += 1
            VLoad = lis
            lis = []
            tic = 0
            for vp in ILoad:
                lis.append(float(ILoad[tic]))
                tic += 1
            ILoad = lis
            lis = []
            tic = 0
            for vp in PLoad:
                lis.append(float(PLoad[tic]))
                tic += 1
            PLoad = lis
            print(VPanel)
            print(IPanel)
            print(PPanel)
            print(VBatt)
            print(IBatt)
            print(PBatt)
            print(VLoad)
            print(ILoad)
            print(PLoad)

            kop = '_'
            for ty in VPanel:
                a1 += str(ty)
                a1 += '_'
            a1.rstrip(kop)
            for ty in IPanel:
                a2 += str(ty)
                a2 += '_'
            a2.rstrip(kop)
            for ty in PPanel:
                a3 += str(ty)
                a3 += '_'
            a3.rstrip(kop)
            for ty in VBatt:
                a4 += str(ty)
                a4 += '_'
            a4.rstrip(kop)
            for ty in IBatt:
                a5 += str(ty)
                a5 += '_'
            a5.rstrip(kop)
            for ty in PBatt:
                a6 += str(ty)
                a6 += '_'
            a6.rstrip(kop)
            for ty in VLoad:
                a7 += str(ty)
                a7 += '_'
            a7.rstrip(kop)
            for ty in ILoad:
                a8 += str(ty)
                a8 += '_'
            a8.rstrip(kop)
            for ty in PLoad:
                a9 += str(ty)
                a9 += '_'
            a9.rstrip(kop)

           
            ref = db.reference('/01/' + str(dart) + ':')  # Замена для базы данных

            # Сохранение данных
            data = {
                'VPanel': a1,
                'IPanel': a2,
                'PPanel': a3,
                'VBatt': a4,
                'IBatt': a5,
                'PBatt': a6,
                'VLoad': a7,
                'ILoad': a8,
                'PLoad': a9
            }

            # Save the data to the database
            ref.set(data)
            return VPanel, IPanel, PPanel, VBatt, IBatt, PBatt, VLoad, ILoad, PLoad, date, dart
        else:
            VPanel = 0
            IPanel = 0
            PPanel = 0
            VBatt = 0
            IBatt = 0
            PBatt = 0
            VLoad = 0
            ILoad = 0
            PLoad = 0
            return VPanel, IPanel, PPanel, VBatt, IBatt, PBatt, VLoad, ILoad, PLoad, date, dart

def button_clicked(date_variable):
    # Создание нового окна
    full_variable = firebase(date_variable)
    if full_variable[0] == 0:
        graph_window = tk.Toplevel(root)
        graph_window.title("Графики")
        label = ttk.Label(graph_window, text="Данные отсутствуют", background="#FFCDD2", foreground="#B71C1C",
                          padding=8, width=100, anchor=CENTER)
        label.pack(expand=True)
    else:
        VPanel = full_variable[0]
        IPanel = full_variable[1]
        PPanel = full_variable[2]
        VBatt = full_variable[3]
        IBatt = full_variable[4]
        PBatt = full_variable[5]
        VLoad = full_variable[6]
        ILoad = full_variable[7]
        PLoad = full_variable[8]
        date = full_variable[9]
        dart = full_variable[10]
        skob = [date, VPanel, IPanel, PPanel, VBatt, IBatt, PBatt, VLoad, ILoad, PLoad]
        letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

        def excel_file():
            pos_variable = 0
            filename = r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей"
            if os.path.exists(filename):
                print("Указанный файл существует")
            else:
                print("Файл не существует")
                os.mkdir(r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей")
                # Создание вкладок
            if int(dart[0:4]) == 2022:
                filename = r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022"
                if os.path.exists(filename):
                    print("Указанный файл существует")
                else:
                    print("Файл не существует")
                    os.mkdir(r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022")
                    # Создание вкладок
            elif int(dart[0:4]) == 2023:
                filename = r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023"
                if os.path.exists(filename):
                    print("Указанный файл существует")
                else:
                    print("Файл не существует")
                    os.mkdir(r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023")
                    # Создание вкладок
            else:
                filename = r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024"
                if os.path.exists(filename):
                    print("Указанный файл существует")
                else:
                    print("Файл не существует")
                    os.mkdir(r"C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024")
                    # Создание вкладок


            # Путь к файлу Excel
            if int(dart[0:4]) == 2022:
                if int(dart[5:6]) == 1:
                    if ':' in dart[5:7]:
                        file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\1.xlsx'
                        pos_variable = 1
                    else:
                        if int(dart[5:7]) == 10:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\10.xlsx'
                            pos_variable = 2
                        elif int(dart[5:7]) == 11:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\11.xlsx'
                            pos_variable = 2
                        else:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\12.xlsx'
                            pos_variable = 2
                elif int(dart[5:6]) == 2:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\2.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 3:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\3.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 4:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\4.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 5:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\5.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 6:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\6.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 7:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\7.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 8:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\8.xlsx'
                    pos_variable = 1
                else:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2022\9.xlsx'
                    pos_variable = 1

            elif int(dart[0:4]) == 2023:
                if int(dart[5:6]) == 1:
                    if ':' in dart[5:7]:
                        file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\1.xlsx'
                        pos_variable = 1
                    else:
                        if int(dart[5:7]) == 10:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\10.xlsx'
                            pos_variable = 2
                        elif int(dart[5:7]) == 11:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\11.xlsx'
                            pos_variable = 2
                        else:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\12.xlsx'
                            pos_variable = 2
                elif int(dart[5:6]) == 2:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\2.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 3:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\3.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 4:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\4.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 5:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\5.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 6:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\6.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 7:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\7.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 8:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\8.xlsx'
                    pos_variable = 1
                else:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2023\9.xlsx'
                    pos_variable = 1
            else:
                if int(dart[5:6]) == 1:
                    if ':' in dart[5:7]:
                        file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\1.xlsx'
                        pos_variable = 1
                    else:
                        if int(dart[5:7]) == 10:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\10.xlsx'
                            pos_variable = 2
                        elif int(dart[5:7]) == 11:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\11.xlsx'
                            pos_variable = 2
                        else:
                            file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\12.xlsx'
                            pos_variable = 2
                elif int(dart[5:6]) == 2:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\2.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 3:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\3.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 4:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\4.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 5:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\5.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 6:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\6.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 7:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\7.xlsx'
                    pos_variable = 1
                elif int(dart[5:6]) == 8:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\8.xlsx'
                    pos_variable = 1
                else:
                    file_path = r'C:\Users\Lenovo\Desktop\Мониторинг Солнечных Панелей\2024\9.xlsx'
                    pos_variable = 1
            if pos_variable == 1:
                if ':' in dart[7:9]:
                    sheet_name = str(dart[7:8])  # Название вкладки
                else:
                    sheet_name = str(dart[7:9])  # Название вкладки
            else:
                if ':' in dart[8:10]:
                    sheet_name = str(dart[8:9])  # Название вкладки
                else:
                    sheet_name = str(dart[8:10])  # Название вкладки

            if not os.path.exists(file_path):
                # Создаем новый файл Excel, если он отсутствует
                wb = Workbook()
                sheet = wb.active
                sheet.title = sheet_name
            else:
                # Загружаем существующий файл Excel
                wb = load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    # Если вкладка существует, выбираем ее
                    sheet = wb[sheet_name]
                else:
                    # Если вкладка не существует, создаем новую
                    sheet = wb.create_sheet(title=sheet_name)

            # Добавляем данные в вкладку
            base_d = ['Время', 'Напряжение солнечной панели', 'Ток солнечной панели', 'Мощность солнечной панели',
                      'Напряжение с аккумулятора', 'Ток с аккумулятора', 'Мощность с аккумулятора',
                      'Напряжение с нагрузки', 'Ток с нагрузки', 'Мощностей с нагрузки']
            for i, value in enumerate(base_d):
                column = chr(ord('A') + i)  # Преобразуем индекс в букву столбца
                cell = f"{column}1"  # Записываем данные в первую строку каждого столбца
                sheet[cell] = value
            for ind in range(10):
                kul = skob[ind]
                let = letters[ind]
                for i, value in enumerate(kul, start=1):
                    column = let
                    column += str(i)
                    sheet[column] = value

            # Сохраняем файл
            wb.save(file_path)


        graph_window = tk.Toplevel(root)
        graph_window.title("Графики")

        # Создание вкладок
        tab_control = ttk.Notebook(graph_window)

        # Вкладка 1
        tab1 = ttk.Frame(tab_control)
        tab_control.add(tab1, text='Напряжение')

        # Вкладка 2
        tab2 = ttk.Frame(tab_control)
        tab_control.add(tab2, text='Ток')

        # Вкладка 3
        tab3 = ttk.Frame(tab_control)
        tab_control.add(tab3, text='Мощность')

        tab_control.pack(expand=1, fill='both')

        # Создание Frame на каждой вкладке
        frame1 = tk.Frame(tab1)
        frame1.pack(fill=tk.BOTH, expand=True)

        frame2 = tk.Frame(tab2)
        frame2.pack(fill=tk.BOTH, expand=True)

        frame3 = tk.Frame(tab3)
        frame3.pack(fill=tk.BOTH, expand=True)

        # Создание графика 1
        fig1 = Figure(figsize=(5, 4), dpi=100)
        ax1 = fig1.add_subplot(111)
        ax1.plot(date, VPanel, label="Напряжение панели", color="red")
        ax1.plot(date, VBatt, label="Напряжение аккумулятора", color="yellow")
        ax1.plot(date, VLoad, label="Напряжение нагрузки", color="green")
        ax1.legend()

        # Вставка графика 1 во Frame
        canvas1 = FigureCanvasTkAgg(fig1, master=frame1)
        canvas1.draw()
        canvas1.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Создание графика 2
        fig2 = Figure(figsize=(5, 4), dpi=100)
        ax2 = fig2.add_subplot(111)
        ax2.plot(date, IPanel, label="Ток панели", color="red")
        ax2.plot(date, IBatt, label="Ток аккумулятора", color="yellow")
        ax2.plot(date, ILoad, label="Ток нагрузки", color="green")
        ax2.legend()

        # Вставка графика 2 во Frame
        canvas2 = FigureCanvasTkAgg(fig2, master=frame2)
        canvas2.draw()
        canvas2.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Создание графика 3
        fig3 = Figure(figsize=(5, 4), dpi=100)
        ax3 = fig3.add_subplot(111)
        ax3.plot(date, PPanel, label="Мощность панели", color="red")
        ax3.plot(date, PBatt, label="Мощность аккумулятора", color="yellow")
        ax3.plot(date, PLoad, label="Мощность нагрузки", color="green")
        ax3.legend()

        # Вставка графика 3 во Frame
        canvas3 = FigureCanvasTkAgg(fig3, master=frame3)
        canvas3.draw()
        canvas3.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        zoom_button = tk.Button(graph_window, text="Создать Excel файл", command=excel_file)
        zoom_button.pack()



def retrieve_details(event):
    item = event.widget.focus()
    popup = tk.Toplevel(root)
    popup.title(event.widget.item(item)['text'])
    popup.geometry("2500x1500")
    popup_label = ttk.Label(popup, text=event)
    popup_label.pack(pady=100)
    close_button = ttk.Button(popup, text="Закрыть", command=popup.destroy)
    close_button.pack()


def open_new_window1(variable1):
    new_window = tk.Toplevel(root)
    if variable1 == 1:
        btn_1 = tk.Button(new_window, text="Январь", command=open_new_january1)
        btn_1.grid(row=0, column=0, sticky=W, pady=10, padx=10)
        btn_2 = tk.Button(new_window, text="Февраль", command=open_new_february1)
        btn_2.grid(row=0, column=1, sticky=W, pady=10, padx=10)
        btn_3 = tk.Button(new_window, text="Март", command=open_new_march1)
        btn_3.grid(row=0, column=2, sticky=W, pady=10, padx=10)
        btn_4 = tk.Button(new_window, text="Апрель", command=open_new_april1)
        btn_4.grid(row=1, column=0, sticky=W, pady=10, padx=10)
        btn_5 = tk.Button(new_window, text="Май", command=open_new_may1)
        btn_5.grid(row=1, column=1, sticky=W, pady=10, padx=10)
        btn_6 = tk.Button(new_window, text="Июнь", command=open_new_june1)
        btn_6.grid(row=1, column=2, sticky=W, pady=10, padx=10)
        btn_7 = tk.Button(new_window, text="Июль", command=open_new_july1)
        btn_7.grid(row=2, column=0, sticky=W, pady=10, padx=10)
        btn_8 = tk.Button(new_window, text="Август", command=open_new_august1)
        btn_8.grid(row=2, column=1, sticky=W, pady=10, padx=10)
        btn_9 = tk.Button(new_window, text="Сентябрь", command=open_new_september1)
        btn_9.grid(row=2, column=2, sticky=W, pady=10, padx=10)
        btn_10 = tk.Button(new_window, text="Октрябрь", command=open_new_october1)
        btn_10.grid(row=3, column=0, sticky=W, pady=10, padx=10)
        btn_11 = tk.Button(new_window, text="Ноябрь", command=open_new_november1)
        btn_11.grid(row=3, column=1, sticky=W, pady=10, padx=10)
        btn_12 = tk.Button(new_window, text="Декабрь", command=open_new_december1)
        btn_12.grid(row=3, column=2, sticky=W, pady=10, padx=10)
    elif variable1 == 2:
        btn_1 = tk.Button(new_window, text="Январь", command=open_new_january2)
        btn_1.grid(row=0, column=0, sticky=W, pady=10, padx=10)
        btn_2 = tk.Button(new_window, text="Февраль", command=open_new_february2)
        btn_2.grid(row=0, column=1, sticky=W, pady=10, padx=10)
        btn_3 = tk.Button(new_window, text="Март", command=open_new_march2)
        btn_3.grid(row=0, column=2, sticky=W, pady=10, padx=10)
        btn_4 = tk.Button(new_window, text="Апрель", command=open_new_april2)
        btn_4.grid(row=1, column=0, sticky=W, pady=10, padx=10)
        btn_5 = tk.Button(new_window, text="Май", command=open_new_may2)
        btn_5.grid(row=1, column=1, sticky=W, pady=10, padx=10)
        btn_6 = tk.Button(new_window, text="Июнь", command=open_new_june2)
        btn_6.grid(row=1, column=2, sticky=W, pady=10, padx=10)
        btn_7 = tk.Button(new_window, text="Июль", command=open_new_july2)
        btn_7.grid(row=2, column=0, sticky=W, pady=10, padx=10)
        btn_8 = tk.Button(new_window, text="Август", command=open_new_august2)
        btn_8.grid(row=2, column=1, sticky=W, pady=10, padx=10)
        btn_9 = tk.Button(new_window, text="Сентябрь", command=open_new_september2)
        btn_9.grid(row=2, column=2, sticky=W, pady=10, padx=10)
        btn_10 = tk.Button(new_window, text="Октрябрь", command=open_new_october2)
        btn_10.grid(row=3, column=0, sticky=W, pady=10, padx=10)
        btn_11 = tk.Button(new_window, text="Ноябрь", command=open_new_november2)
        btn_11.grid(row=3, column=1, sticky=W, pady=10, padx=10)
        btn_12 = tk.Button(new_window, text="Декабрь", command=open_new_december2)
        btn_12.grid(row=3, column=2, sticky=W, pady=10, padx=10)
    else:
        btn_1 = tk.Button(new_window, text="Январь", command=open_new_january3)
        btn_1.grid(row=0, column=0, sticky=W, pady=10, padx=10)
        btn_2 = tk.Button(new_window, text="Февраль", command=open_new_february3)
        btn_2.grid(row=0, column=1, sticky=W, pady=10, padx=10)
        btn_3 = tk.Button(new_window, text="Март", command=open_new_march3)
        btn_3.grid(row=0, column=2, sticky=W, pady=10, padx=10)
        btn_4 = tk.Button(new_window, text="Апрель", command=open_new_april3)
        btn_4.grid(row=1, column=0, sticky=W, pady=10, padx=10)
        btn_5 = tk.Button(new_window, text="Май", command=open_new_may3)
        btn_5.grid(row=1, column=1, sticky=W, pady=10, padx=10)
        btn_6 = tk.Button(new_window, text="Июнь", command=open_new_june3)
        btn_6.grid(row=1, column=2, sticky=W, pady=10, padx=10)
        btn_7 = tk.Button(new_window, text="Июль", command=open_new_july3)
        btn_7.grid(row=2, column=0, sticky=W, pady=10, padx=10)
        btn_8 = tk.Button(new_window, text="Август", command=open_new_august3)
        btn_8.grid(row=2, column=1, sticky=W, pady=10, padx=10)
        btn_9 = tk.Button(new_window, text="Сентябрь", command=open_new_september3)
        btn_9.grid(row=2, column=2, sticky=W, pady=10, padx=10)
        btn_10 = tk.Button(new_window, text="Октрябрь", command=open_new_october3)
        btn_10.grid(row=3, column=0, sticky=W, pady=10, padx=10)
        btn_11 = tk.Button(new_window, text="Ноябрь", command=open_new_november3)
        btn_11.grid(row=3, column=1, sticky=W, pady=10, padx=10)
        btn_12 = tk.Button(new_window, text="Декабрь", command=open_new_december3)
        btn_12.grid(row=3, column=2, sticky=W, pady=10, padx=10)






def schedule1():
    open_new_window1(1)


def schedule2():
    open_new_window1(2)


def schedule3():
    open_new_window1(3)


def open_new_january1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:01:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:1:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_february1():
    new_window = tk.Toplevel()
    for i in range(28):
        button = tk.Button(new_window, text="2022:02:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:2:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_march1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:03:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:3:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_april1():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2022:04:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:4:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_may1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:05:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:5:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_june1():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2022:06:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:6:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_july1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:07:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:7:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_august1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:08:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:8:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_september1():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2022:09:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:9:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_october1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:10:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:10:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_november1():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2022:11:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:11:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_december1():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2022:12:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2022:12:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_january2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:01:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:1:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_february2():
    new_window = tk.Toplevel()
    for i in range(28):
        button = tk.Button(new_window, text="2023:02:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:2:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_march2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:03:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:3:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_april2():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2023:04:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:4:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_may2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:05:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:5:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_june2():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2023:06:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:6:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_july2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:07:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:7:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_august2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:08:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:8:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_september2():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2023:09:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:9:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_october2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:10:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:10:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_november2():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2023:11:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:11:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_december2():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2023:12:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2023:12:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_january3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:01:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:1:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_february3():
    new_window = tk.Toplevel()
    for i in range(28):
        button = tk.Button(new_window, text="2024:02:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:2:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_march3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:03:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:3:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_april3():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2024:04:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:4:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_may3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:05:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:5:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_june3():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2024:06:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:6:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_july3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:07:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:7:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_august3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:08:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:8:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_september3():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2024:09:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:9:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_october3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:10:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:10:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_november3():
    new_window = tk.Toplevel()
    for i in range(30):
        button = tk.Button(new_window, text="2024:11:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:11:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def open_new_december3():
    new_window = tk.Toplevel()
    for i in range(31):
        button = tk.Button(new_window, text="2024:12:{}".format(i + 1), command=lambda num=i + 1: button_clicked("2024:12:" + str(num)))
        button.grid(column=i % 5, row=i // 5, pady=10, padx=10)


def calculate2(variable4):
    r = 1


def open_new_window3(variable5):
    new_window = tk.Toplevel()
    label = tk.Label(new_window, text=str(1))
    label.pack()


def calculate3(variable6):
    open_new_window3(float(1))


root = tk.Tk()
root.title("Моониторинг Соолнечной Панели")
root.geometry("300x200")

tab_control = ttk.Notebook(root)
tab1 = ttk.Frame(tab_control)
tab2 = ttk.Frame(tab_control)
tab3 = ttk.Frame(tab_control)
tab4 = ttk.Frame(tab_control)
tab_control.add(tab1, text='1-ая панель')
tab_control.add(tab2, text='2-ая панель')
tab_control.add(tab3, text='3-ая панель')
tab_control.add(tab4, text='4-ая панель')

tab_control.pack(expand=1, fill='both')

btn1 = tk.Button(tab1, text="                    2022                    ", command=schedule1)
btn1.grid(row=0, column=0, sticky=W, pady=10, padx=70)
btn2 = tk.Button(tab1, text="                    2023                    ", command=schedule2)
btn2.grid(row=1, column=0, sticky=W, pady=20, padx=70)
btn3 = tk.Button(tab1, text="                    2024                    ", command=schedule3)
btn3.grid(row=2, column=0, sticky=W, pady=20, padx=70)

root.mainloop()
